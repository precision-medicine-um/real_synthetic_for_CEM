import os
import random
import shutil
import openpyxl
from utils_readwrite import read_dicom
from process_masks import find_mask_region

# Mapping of column names to their respective Excel columns
columnlist = {
    "patient": "A",
    "laterality": "B",
    "CC": "G",
    "MLO": "H"
}

def open_workbook(filepath, filename):
    """Open an Excel workbook and return it."""

    wb = openpyxl.load_workbook(os.path.join(filepath, filename))

    return wb


def close_workbook(filepath, filename, wb):
    """Save and close the Excel workbook."""

    wb.save(os.path.join(filepath, filename))
    wb.close()


def open_worksheet(filepath, filename, sheetname):
    """Open a specific worksheet in the workbook."""

    wb = open_workbook(filepath, filename)
    ws = wb[sheetname]

    return ws


def get_list_worksheet(ws, listname):
    """Retrieve a list from a specific column in the worksheet"""

    return [cell.value for cell in ws[columnlist[listname]][1:]]


def split_cases(filepath, filename, sheetname, sheetout1, sheetout2, percentage=0.5):
    """Split cases from the input sheet into two output sheets based on a percentage."""

    bookin = open_workbook(filepath, filename)
    sheetin = bookin[sheetname]

    # Get lists of patients and their laterality
    patient_list = get_list_worksheet(sheetin, 'patient')
    laterality_list = get_list_worksheet(sheetin, 'laterality')

    # Randomly select a percentage of cases for the first list
    patient1_list, laterality1_list = zip(*random.sample(list(zip(patient_list, laterality_list)), int(len(patient_list) * percentage)))
    patient1_list, laterality1_list = zip(*sorted(zip(patient1_list, laterality1_list)))

    # Create a copy for the second list and remove selected cases
    patient2_list = [p for p in patient_list if p not in patient1_list]
    laterality2_list = [l for l in laterality_list if l not in laterality1_list] 
    
    # Create the first output sheet with selected cases
    ws1 = bookin.create_sheet(sheetout1)
    for idx in range(len(patient1_list)):
        for row in sheetin.iter_rows(min_row=1):
            if row[0].value == patient1_list[idx] and row[1].value == laterality1_list[idx]:
                ws1.append((cell.value for cell in row))

    # Create the second output sheet with other cases
    ws2 = bookin.create_sheet(sheetout2)
    for idx in range(len(patient2_list)):
        for row in sheetin.iter_rows(min_row=1):
            if row[0].value == patient2_list[idx] and row[1].value == laterality2_list[idx]:
                ws2.append((cell.value for cell in row))

    close_workbook(filepath, filename, bookin)


def create_bbox_excel(excelpath, excelname, excelsheet,
                      simulpath, imagepath, savepath,
                      bboxpath, bboxname, bboxsheet,
                      origrecombined=True, minnum=0):
    """Create an Excel file with simulated cases and their bounding boxes."""
    
    ws_input = open_worksheet(excelpath, excelname, excelsheet)
    wb_bbox = open_workbook(bboxpath, bboxname)
    ws_bbox = wb_bbox[bboxsheet]
    
    last_row = ws_bbox.max_row

    # Loop over all input cases
    for idx in range(1, len(ws_input['A'])):
        if ws_input['E'][idx].value is not None and minnum < ws_input['C'][idx].value:
            last_row += 1
            
            # Extract parameters for the case
            patientname = ws_input['A'][idx].value
            laterality = ws_input['B'][idx].value
            clusternum = ws_input['C'][idx].value
            malignancy = ws_input['D'][idx].value
            print(idx, patientname, laterality, clusternum, malignancy)

            # Create directories for the simulated images
            cc_dir = os.path.join(savepath, patientname, laterality + '_CC')
            mlo_dir = os.path.join(savepath, patientname, laterality + '_MLO')
            os.makedirs(cc_dir, exist_ok=True)
            os.makedirs(mlo_dir, exist_ok=True)

            # Copy simulated images to output directories
            for casedir in os.listdir(simulpath):
                if 'voxel' in casedir:
                    for imgname in os.listdir(os.path.join(simulpath, casedir, 'Simulated')):
                        if patientname in imgname:
                            if 'CC' in imgname:
                                shutil.copyfile(os.path.join(simulpath, casedir, 'Simulated', imgname), os.path.join(cc_dir, imgname))
                                ws_bbox.cell(row=last_row, column=4).value = 1
                            elif 'MLO' in imgname:
                                shutil.copyfile(os.path.join(simulpath, casedir, 'Simulated', imgname), os.path.join(mlo_dir, imgname))
                                ws_bbox.cell(row=last_row, column=5).value = 1

            # Add case information to the bounding box sheet
            ws_bbox.cell(row=last_row, column=1).value = patientname
            ws_bbox.cell(row=last_row, column=2).value = laterality
            ws_bbox.cell(row=last_row, column=3).value = laterality
            ws_bbox.cell(row=last_row, column=6).value = 'simulated cluster'
            ws_bbox.cell(row=last_row, column=7).value = malignancy
            ws_bbox.cell(row=last_row, column=8).value = malignancy
            
            # Optionally copy the original recombined image
            if origrecombined:
                real_cc_dir = os.path.join(imagepath, patientname, laterality + '_CC') 
                real_mlo_dir = os.path.join(imagepath, patientname, laterality + '_MLO')
                for imgname in os.listdir(real_cc_dir):
                    if 'REC' in imgname:
                        shutil.copyfile(os.path.join(real_cc_dir, imgname), os.path.join(cc_dir, imgname))
                for imgname in os.listdir(real_mlo_dir):
                    if 'REC' in imgname:
                        shutil.copyfile(os.path.join(real_mlo_dir, imgname), os.path.join(mlo_dir, imgname))

            # Find and store the mask region for CC
            for imgname in os.listdir(cc_dir):
                if 'mask' in imgname or 'MASK' in imgname:
                    mask_cc = read_dicom(cc_dir, imgname, force=True)
                    mask_array_cc = mask_cc.pixel_array 

            xmin, xmax, ymin, ymax = find_mask_region(mask_array_cc)
            if xmin[0] > 0:
                ws_bbox.cell(row=last_row, column=9).value = '{},{},{},{}'.format(xmin, ymin, xmax, ymax)

            # Find and store the mask region for MLO
            for imgname in os.listdir(mlo_dir):
                if 'mask' in imgname or 'MASK' in imgname:
                    mask_mlo = read_dicom(mlo_dir, imgname, force=True)
                    mask_array_mlo = mask_mlo.pixel_array

            xmin, xmax, ymin, ymax = find_mask_region(mask_array_mlo)
            if xmin[0] > 0:
                ws_bbox.cell(row=last_row, column=10).value = '{},{},{},{}'.format(xmin, ymin, xmax, ymax)

    close_workbook(bboxpath, bboxname, wb_bbox)
    return wb_bbox


